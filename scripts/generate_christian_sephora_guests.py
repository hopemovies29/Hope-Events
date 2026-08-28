"""Build one QR and invitation folder per guest from the Christian/Sephora workbook."""

from __future__ import annotations

import json
import os
import re
import unicodedata
from difflib import SequenceMatcher
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT.parent / "Mes invités.xlsx"
EVENT_ROOT = ROOT / "public" / "couple-christian-sephora"
TEMPLATE = EVENT_ROOT / "Table Clou de girofle" / "couple-palama"
PUBLIC_ROOT = "https://hope-events.vercel.app"
EVENT = {
    "id": "christian-sephora-palama-2026",
    "slug": "christian-sephora-palama-2026",
    "accessKey": "HE-CSM-2026",
    "coupleNames": "Christian Lengbe et Sephora Malanda",
    "dateIso": "2026-09-12T19:00:00+01:00",
    "dateLabel": "Samedi 12 septembre 2026",
    "timeLabel": "Benediction 11h00 - Soiree 19h00",
    "venueName": "Salle Chapiteau La Colombe",
    "venueAddress": "Av. Macampagne N°31 — Arrêt Quado, Coin Bocage",
    "mapUrl": "https://maps.google.com/?q=Salle+Chapiteau+La+Colombe+Macampagne+31+Ngaliema+Kinshasa",
    "eventPhrase": "Une belle histoire se poursuit, soyez a nos cotes pour ecrire avec nous le prochain chapitre de notre amour.",
    "preferences": {
        "beers": ["Heineken", "Likofi", "Castel", "Tembo", "Beaufort", "Primus", "Turbo King", "Nkoyi Grand", "33 Export", "Savanna", "Bavaria"],
        "wine": ["Vin rouge", "Whisky"],
        "soft": ["Coca-Cola", "Fanta", "Sprite", "Vital'O", "Maltina", "Énergie Malt", "XXL", "Top", "Eau"],
    },
}


def text(value: object) -> str:
    return str(value or "").strip()


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    cleaned = re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")
    return cleaned or "invite"


def canonical_table(value: str) -> str:
    aliases = {"clou de giroffle": "Table Clou de girofle"}
    return aliases.get(value.lower(), value)


def table_token_slug(table_name: str) -> str:
    if table_name == "Table Clou de girofle":
        return "clou-de-girofle"
    return slug(table_name)


def usable_guest(value: object) -> bool:
    candidate = text(value)
    if not candidate or candidate.upper() in {"TOTAL", "TABLES", "SEPHORA"}:
        return False

    try:
        float(candidate.replace(",", "."))
        return False
    except ValueError:
        return True


def table_layout(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int]:
    """Return the header row and first table column for each supplied workbook layout."""
    normalized_title = slug(sheet.title)

    if normalized_title == "lengbe-christian":
        return 5, 2

    if normalized_title == "malanda-sephora":
        return 3, 4

    # Keep the earlier workbook layout supported for a safe regeneration fallback.
    if sheet.title == "Christian":
        header_row = next(
            row
            for row in range(1, sheet.max_row + 1)
            if text(sheet.cell(row, 1).value).lower() == "tables"
        )
        return header_row, 2

    title_row = next(
        row
        for row in range(1, sheet.max_row + 1)
        if text(sheet.cell(row, 2).value).upper() == "SEPHORA"
    )
    return title_row, 3


def table_columns(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[int, str]:
    header_row, start_column = table_layout(sheet)

    columns = {}
    duplicate_names: defaultdict[str, int] = defaultdict(int)
    for column in range(start_column, sheet.max_column + 1):
        name = text(sheet.cell(header_row, column).value)
        if name:
            canonical_name = canonical_table(name)
            duplicate_names[canonical_name] += 1
            suffix = "" if duplicate_names[canonical_name] == 1 else f" {duplicate_names[canonical_name]}"
            columns[column] = f"{canonical_name}{suffix}"

    return columns


def previous_directory() -> dict[str, dict]:
    """Read the previous directory before replacing it during generation."""
    directory = Path(
        os.environ.get(
            "CHRISTIAN_SEPHORA_PREVIOUS_DIRECTORY",
            ROOT / "data" / "christian-sephora-guests.js",
        )
    )

    if not directory.exists():
        return {"invitations": {}, "routes": {}}

    source = directory.read_text(encoding="utf-8")
    match = re.search(r"const invitations = (\{.*?\});\nconst routes", source, re.DOTALL)

    if not match:
        return {"invitations": {}, "routes": {}}

    invitations = json.loads(match.group(1))
    routes_match = re.search(r"const routes = (\{.*?\});\n\nmodule.exports", source, re.DOTALL)
    routes = json.loads(routes_match.group(1)) if routes_match else {}
    return {"invitations": invitations, "routes": routes}


def previous_tokens_by_guest(invitations: dict[str, dict]) -> dict[str, list[str]]:
    """Reuse old tokens so QR cards already sent to guests keep working."""
    tokens: defaultdict[str, list[str]] = defaultdict(list)
    for token, invitation in invitations.items():
        guest_name = text(invitation.get("guestName"))
        if guest_name:
            tokens[slug(guest_name)].append(token)
    return tokens


def guest_identity(value: str) -> str:
    """Normalize titles and spacing so renamed spreadsheet entries retain their QR token."""
    return (
        slug(value)
        .replace("couple", "")
        .replace("cpl", "")
        .replace("mme", "")
        .replace("mr", "")
        .replace("dr", "")
        .replace("honorable", "")
        .replace("onorable", "")
    )


def find_previous_token(
    guest_name: str,
    previous_invitations: dict[str, dict],
    previous_tokens: dict[str, list[str]],
    used_tokens: set[str],
) -> str:
    exact_tokens = previous_tokens.get(slug(guest_name), [])
    while exact_tokens:
        token = exact_tokens.pop(0)
        if token not in used_tokens:
            return token

    identity = guest_identity(guest_name)
    candidates = [
        (token, invitation)
        for token, invitation in previous_invitations.items()
        if token not in used_tokens
    ]

    for token, invitation in candidates:
        previous_identity = guest_identity(text(invitation.get("guestName")))
        if identity and (identity == previous_identity or identity in previous_identity or previous_identity in identity):
            return token

    scored = [
        (
            SequenceMatcher(
                None, identity, guest_identity(text(invitation.get("guestName")))
            ).ratio(),
            token,
        )
        for token, invitation in candidates
    ]
    score, token = max(scored, default=(0, ""))
    return token if score >= 0.82 else ""


def extract_guests(
    workbook: openpyxl.Workbook,
    previous_invitations: dict[str, dict],
    previous_tokens: dict[str, list[str]],
) -> list[dict[str, str | int]]:
    records: list[dict[str, str | int]] = []
    used_previous_tokens: set[str] = set()

    for sheet in workbook.worksheets:
        header_row, _ = table_layout(sheet)
        columns = table_columns(sheet)
        first_guest_row = header_row + 1
        used_slugs: defaultdict[str, int] = defaultdict(int)

        for column, table_name in columns.items():
            for row in range(first_guest_row, sheet.max_row + 1):
                guest_name = text(sheet.cell(row, column).value)
                if not usable_guest(guest_name):
                    continue
                base_slug = f"{table_token_slug(table_name)}-{slug(guest_name)}"
                used_slugs[base_slug] += 1
                suffix = "" if used_slugs[base_slug] == 1 else f"-{used_slugs[base_slug]}"
                token = find_previous_token(
                    guest_name,
                    previous_invitations,
                    previous_tokens,
                    used_previous_tokens,
                )
                if token:
                    used_previous_tokens.add(token)
                else:
                    token = f"{base_slug}{suffix}"
                records.append(
                    {
                        "token": token,
                        "guestName": guest_name,
                        "tableName": table_name,
                        "tableSlug": table_token_slug(table_name),
                        "folderSlug": f"{slug(guest_name)}{suffix}",
                    }
                )

    return records


def merge_legacy_guests(
    guests: list[dict[str, str | int]], previous: dict[str, dict]
) -> list[dict[str, str | int]]:
    """Keep prior invite records that are not present in a replacement workbook."""
    active_tokens = {str(guest["token"]) for guest in guests}

    for token, invitation in previous["invitations"].items():
        if token in active_tokens:
            continue

        route = previous["routes"].get(token, {})
        invitation_path = str(route.get("invitationPagePath", ""))
        path_parts = invitation_path.replace("\\", "/").split("/")
        folder_slug = path_parts[-2] if len(path_parts) >= 2 else slug(text(invitation.get("guestName")))
        guests.append(
            {
                "token": token,
                "guestName": text(invitation.get("guestName")),
                "tableName": text(invitation.get("tableName")) or "Table a attribuer",
                "tableSlug": text(invitation.get("tableSlug")) or "table-a-attribuer",
                "folderSlug": folder_slug,
            }
        )

    return guests


def public_paths(table_name: str, guest_slug: str) -> tuple[str, str]:
    table_path = quote(table_name, safe="")
    invitation = f"/couple-christian-sephora/{table_path}/{guest_slug}/invitation"
    qr = f"/couple-christian-sephora/{table_path}/{guest_slug}/qr-code-{guest_slug}"
    return invitation, qr


def write_guest_pages(guest: dict[str, str | int]) -> dict[str, str]:
    table_name = str(guest["tableName"])
    guest_name = str(guest["guestName"])
    token = str(guest["token"])
    guest_slug = str(guest.get("folderSlug") or slug(guest_name))
    folder = EVENT_ROOT / table_name / guest_slug
    folder.mkdir(parents=True, exist_ok=True)
    invitation_public_path, qr_public_path = public_paths(table_name, guest_slug)
    poster_filename = "Invitation_Couple_Palama_Table_Clou_de_girofle.png"
    poster_path = (
        f"./{poster_filename}"
        if token == "clou-de-girofle-couple-palama"
        else f"../../Table Clou de girofle/couple-palama/{poster_filename}"
    )

    replacements = [
        (
            "/couple-christian-sephora/Table%20Clou%20de%20girofle/couple-palama/invitation",
            invitation_public_path,
        ),
        (
            "/couple-christian-sephora/Table%20Clou%20de%20girofle/couple-palama/qr-code-couple-palama",
            qr_public_path,
        ),
        ("clou-de-girofle-couple-palama", token),
        ("Couple Palama", guest_name),
        ("Table Clou de girofle", table_name),
        ("couple-palama", guest_slug),
    ]

    for filename in ("invitation.html", "invitation.js", "style.css", "qr-code-couple-palama.html"):
        source = TEMPLATE / filename
        destination_name = filename.replace("couple-palama", guest_slug)
        destination = folder / destination_name
        if destination == source:
            continue
        content = source.read_text(encoding="utf-8")
        for before, after in replacements:
            content = content.replace(before, after)
        content = content.replace(
            "./Invitation_Couple_Palama_Table_Clou_de_girofle.png",
            poster_path,
        )
        destination.write_text(content, encoding="utf-8")

    return {
        "invitationPagePath": f"./{table_name}/{guest_slug}/invitation.html",
        "publicInvitationPath": invitation_public_path,
        "qrPagePath": f"./{table_name}/{guest_slug}/qr-code-{guest_slug}.html",
        "publicQrPagePath": qr_public_path,
    }


def write_guest_data(guests: list[dict[str, str | int]], routes: dict[str, dict[str, str]]) -> None:
    invitation_data = {}
    for guest in guests:
        token = str(guest["token"])
        invitation_data[token] = {
            **guest,
            "eventId": EVENT["id"],
            "salutation": f"Cher{'' if str(guest['guestName']).lower().startswith('couple') else 'e'} invite",
            "seats": 2 if str(guest["guestName"]).lower().startswith("couple") else 1,
            "personalMessage": "Christian et Sephora seront heureux de partager cette journee de benediction et de joie avec vous.",
        }

    output = f"""// Private invitation directory used only by Vercel API routes.
const event = {json.dumps(EVENT, ensure_ascii=False, indent=2)};
const invitations = {json.dumps(invitation_data, ensure_ascii=False, indent=2)};
const routes = {json.dumps(routes, ensure_ascii=False, indent=2)};

module.exports = {{ event, invitations, routes }};
"""
    (ROOT / "data" / "christian-sephora-guests.js").write_text(output, encoding="utf-8")


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    previous = previous_directory()
    guests = extract_guests(
        workbook,
        previous["invitations"],
        previous_tokens_by_guest(previous["invitations"]),
    )
    if previous["invitations"]:
        # The existing QR list is authoritative; unmatched new rows require a separate QR issue.
        guests = [
            guest
            for guest in guests
            if str(guest["token"]) in previous["invitations"]
        ]
    unique_guests = {}
    for guest in guests:
        unique_guests.setdefault(str(guest["token"]), guest)
    guests = list(unique_guests.values())
    current_tokens = {str(guest["token"]) for guest in guests}
    guests = merge_legacy_guests(guests, previous)
    routes = {
        str(guest["token"]): write_guest_pages(guest)
        for guest in guests
        if str(guest["token"]) in current_tokens
    }
    routes.update(
        {
            token: route
            for token, route in previous["routes"].items()
            if token not in current_tokens
        }
    )
    write_guest_data(guests, routes)
    print(f"Generated {len(guests)} guest folders across {len(set(str(item['tableName']) for item in guests))} tables.")


if __name__ == "__main__":
    main()
